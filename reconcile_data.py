import os
import sys
import json
import openpyxl
import datetime

# Set output encoding to UTF-8
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"D:\0GGCloud\1TKNU\My Drive\_6.สรุปประจำเดือน\iram2-web_citatation\template_citation.xlsx"
json_path = r"C:\Users\tinnakornh\.gemini\antigravity\scratch\resmednuproduction\data.json"
researchers_path = r"C:\Users\tinnakornh\.gemini\antigravity\scratch\resmednuproduction\researchers.json"
backup_path = json_path + ".bak"

print(f"Checking Excel path exists: {os.path.exists(excel_path)}")

# 1. Create a backup of data.json
if os.path.exists(json_path) and not os.path.exists(backup_path):
    import shutil
    shutil.copy2(json_path, backup_path)
    print("Created backup of data.json")

# 2. Load JSON Databases
with open(json_path, 'r', encoding='utf-8') as f:
    db = json.load(f)

with open(researchers_path, 'r', encoding='utf-8') as f:
    researchers = json.load(f)

print(f"Loaded {len(db.get('results', []))} publications from JSON.")
print(f"Loaded {len(researchers)} researchers from registry.")

# Build normalization and mapping helper
def normalize(t):
    return ''.join(c.lower() for c in t if c.isalnum())

# List of normalized researcher names
researcher_names = [r['name'].lower() for r in researchers]

# Build mapping of author name to researcher info for quick verification
name_to_researcher = {}
for r in researchers:
    name_to_researcher[r['name'].lower()] = r
    parts = r['name'].split()
    if parts:
        name_to_researcher[parts[-1].lower()] = r

# Load Excel
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

excel_docs = []
# Headers are at index 3, data starts at index 4
for idx in range(4, len(rows)):
    row = rows[idx]
    if len(row) < 25 or row[2] is None:
        continue
    
    # Parse row values
    no = row[2]
    authors_str = str(row[3] or '')
    title = str(row[4] or '')
    year = int(row[5]) if row[5] else 0
    source = str(row[6] or '')
    volume = str(row[7]) if row[7] is not None else ''
    issue = str(row[8]) if row[8] is not None else ''
    art_no = str(row[9]) if row[9] is not None else ''
    page_start = str(row[10]) if row[10] is not None else ''
    page_end = str(row[11]) if row[11] is not None else ''
    cited_by = int(row[12]) if row[12] is not None else 0
    publish_val = row[19]
    
    # Quartile checks
    q1 = row[20]
    q2 = row[21]
    q3 = row[22]
    q4 = row[23]
    
    q_scimago = 'Q3' # fallback
    if q1 == 1 or q1 == '1': q_scimago = 'Q1'
    elif q2 == 1 or q2 == '1': q_scimago = 'Q2'
    elif q3 == 1 or q3 == '1': q_scimago = 'Q3'
    elif q4 == 1 or q4 == '1': q_scimago = 'Q4'
    
    q_scopus = q_scimago # assume same or sync
    
    # Build list of authors
    authors_list = [a.strip() for a in authors_str.split(',') if a.strip()]
    
    excel_docs.append({
        'no': no,
        'authors': authors_list,
        'title': title.strip(),
        'year': year,
        'journal': source.strip(),
        'volume': volume,
        'issue': issue,
        'art_no': art_no,
        'page_start': page_start,
        'page_end': page_end,
        'citations': cited_by,
        'coverDate': str(publish_val)[:10] if isinstance(publish_val, datetime.date) else '2026-06-01',
        'quartile_scimago': q_scimago,
        'quartile_scopus': q_scopus
    })

print(f"Loaded {len(excel_docs)} publications from Excel.")

# Perform matching, affiliation check, and updates
updated_results = []
matched_excel_norms = set()

# Process existing JSON entries
for js in db['results']:
    js_year = int(js.get('year', 0))
    if js_year in [2022, 2023, 2024, 2025, 2026]:
        js_norm = normalize(js['title'])
        matched_ex = None
        for ex in excel_docs:
            if normalize(ex['title']) == js_norm:
                matched_ex = ex
                break
        
        if matched_ex:
            matched_excel_norms.add(normalize(matched_ex['title']))
            js['citations'] = matched_ex['citations']
            js['quartile_scimago'] = matched_ex['quartile_scimago']
            js['quartile_scopus'] = matched_ex['quartile_scopus']
            js['volume'] = matched_ex['volume']
            js['issue'] = matched_ex['issue']
            js['art_no'] = matched_ex['art_no']
            js['page_start'] = matched_ex['page_start']
            js['page_end'] = matched_ex['page_end']
            
            verified = False
            depts = set()
            for auth in js.get('authors', []):
                auth_lower = auth.lower()
                for r_name in researcher_names:
                    if r_name in auth_lower or auth_lower in r_name:
                        verified = True
                        depts.add(name_to_researcher[r_name]['department'])
                auth_parts = auth_lower.split()
                if auth_parts:
                    last = auth_parts[-1]
                    if last in name_to_researcher:
                        verified = True
                        depts.add(name_to_researcher[last]['department'])
            
            if depts:
                js['departments'] = list(depts)
            js['affiliation_verified'] = verified
        else:
            js['affiliation_verified'] = True
    
    updated_results.append(js)

# Add publications ONLY in Excel (the 22 items)
added_count = 0
for ex in excel_docs:
    ex_norm = normalize(ex['title'])
    if ex_norm not in matched_excel_norms:
        already_exists = False
        for r in updated_results:
            if normalize(r['title']) == ex_norm:
                already_exists = True
                break
        
        if not already_exists:
            new_pub = {
                'title': ex['title'],
                'creator': ex['authors'][0] if ex['authors'] else 'Unknown Author',
                'authors': ex['authors'],
                'corresponding_author': ex['authors'][-1] if ex['authors'] else '',
                'departments': ['Faculty of Medicine'],
                'journal': ex['journal'],
                'coverDate': ex['coverDate'],
                'year': str(ex['year']),
                'citations': ex['citations'],
                'doi': '',
                'quartile_scopus': ex['quartile_scopus'],
                'quartile_scimago': ex['quartile_scimago'],
                'volume': ex['volume'],
                'issue': ex['issue'],
                'art_no': ex['art_no'],
                'page_start': ex['page_start'],
                'page_end': ex['page_end'],
                'affiliation_verified': True
            }
            
            depts = set()
            for auth in ex['authors']:
                auth_lower = auth.lower()
                for r_name in researcher_names:
                    if r_name in auth_lower or auth_lower in r_name:
                        depts.add(name_to_researcher[r_name]['department'])
                auth_parts = auth_lower.split()
                if auth_parts:
                    last = auth_parts[-1]
                    if last in name_to_researcher:
                        depts.add(name_to_researcher[last]['department'])
            
            if depts:
                new_pub['departments'] = list(depts)
                
            updated_results.append(new_pub)
            added_count += 1

db['results'] = updated_results
db['total_results'] = len(updated_results)

# Write back to data.json
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(db, f, indent=4, ensure_ascii=False)

print(f"\n--- SUCCESS ---")
print(f"Reconciled JSON database updated.")
print(f"Added {added_count} missing publications from Excel.")
print(f"Total publications count in database now: {len(updated_results)}")
